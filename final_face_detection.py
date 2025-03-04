import cv2
import numpy as np
import faiss
import os
import dlib
import pickle
import torch
from deepface import DeepFace
from ultralytics.utils.ops import non_max_suppression
import onnxruntime as ort
import pandas as pd
from datetime import datetime, timedelta
import time
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Set environment variable
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# Load YOLOv11 ONNX model
onnx_path = "yolov11n-face.onnx"
ort_session = ort.InferenceSession(onnx_path)

# FAISS setup
embedding_dim = 512  # Matches FaceNet512 output
faiss_index = faiss.IndexFlatL2(embedding_dim)
embedding_db = {}

# Load existing embeddings if available
if os.path.exists("face_index.pkl"):
    with open("face_index.pkl", "rb") as f:
        embedding_db, faiss_index = pickle.load(f)

print(f"Total faces in database: {faiss_index.ntotal}")

dlib_detector = dlib.get_frontal_face_detector()
predictor_path = "shape_predictor_68_face_landmarks.dat"  # Path to Dlib's landmark predictor
landmark_predictor = dlib.shape_predictor(predictor_path)

# Tracking dictionary for detection counts and timestamps
detection_tracker = {}  # Format: {name: {"count": int, "times": [datetime]}}

# Dictionary to track names logged per date (key: date_str, value: set of names)
logged_names = {}

def preprocess_onnx(img):
    """Prepares the image for ONNX face detection model"""
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img = cv2.resize(img, (640, 640))
    img = img / 255.0
    img = img.transpose(2, 0, 1)
    img = np.expand_dims(img, axis=0).astype(np.float32)
    return img

def postprocess_onnx(outputs, orig_h, orig_w):
    """Processes ONNX model output and scales detections to original image size"""
    outputs = torch.from_numpy(outputs)
    outputs = non_max_suppression(outputs, 0.5, 0.5)
    detections = []
    scale_h, scale_w = orig_h / 640, orig_w / 640
    for output in outputs:
        if output is not None:
            for *xyxy, conf, cls in output:
                x1, y1, x2, y2 = map(int, xyxy)
                x1, y1, x2, y2 = int(x1,scale_w), int(y1, scale_h), int(x2, scale_w), int(y2, scale_h)
                detections.append([x1, y1, x2, y2, float(conf), int(cls)])
    return detections

def align_face(img):
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    faces = dlib_detector(gray)
    for face in faces:
        landmarks = landmark_predictor(gray, face)
        left_eye = (landmarks.part(36).x, landmarks.part(36).y)
        right_eye = (landmarks.part(45).x, landmarks.part(45).y)
        dx, dy = right_eye[0] - left_eye[0], right_eye[1] - left_eye[1]
        angle = np.degrees(np.arctan2(dy, dx))
        center = ((left_eye[0] + right_eye[0]) // 2, (left_eye[1] + right_eye[1]) // 2)
        rotation_matrix = cv2.getRotationMatrix2D(center, angle, 1)
        aligned = cv2.warpAffine(img, rotation_matrix, (img.shape[1], img.shape[0]))
        return aligned
    return img

def is_face_straight(img):
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    landmarks = landmark_predictor(gray, dlib.rectangle(0, 0, gray.shape[1], gray.shape[0]))
    nose = (landmarks.part(30).x, landmarks.part(30).y)
    left_eye = (landmarks.part(36).x, landmarks.part(36).y)
    right_eye = (landmarks.part(45).x, landmarks.part(45).y)
    left_distance = abs(nose[0] - left_eye[0])
    right_distance = abs(nose[0] - right_eye[0])
    if left_distance / right_distance > 2 or right_distance / left_distance > 2:
        print("Skip the image due to looking sideways -------------------------")
        return False
    return True

def extract_embedding(face):
    """Extracts facial embeddings using FaceNet512"""
    if face is None or face.size == 0:
        print("Empty face crop detected.")
        return None

    try:
        face = cv2.resize(face, (160, 160))  # FaceNet512 expects 160x160 input
        face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        embedding = DeepFace.represent(face, model_name="Facenet512", enforce_detection=False)
        
        if isinstance(embedding, list) and len(embedding) > 0:
            if isinstance(embedding[0], dict):
                embedding_vector = np.array(embedding[0].get("embedding", []), dtype=np.float32)
            else:
                embedding_vector = np.array(embedding, dtype=np.float32)

            if embedding_vector.size == embedding_dim:
                embedding_vector = embedding_vector / np.linalg.norm(embedding_vector)
                return embedding_vector.reshape(1, -1)
    except Exception as e:
        print(f"Embedding extraction failed: {e}")
    return None

def store_embeddings(image_folder):
    """Stores embeddings for images in a given folder"""
    global faiss_index, embedding_db

    if not os.path.exists(image_folder):
        print(f"Folder '{image_folder}' not found.")
        return

    for filename in os.listdir(image_folder):
        img_path = os.path.join(image_folder, filename)
        img = cv2.imread(img_path)

        if img is None:
            print(f"Could not read {img_path}, skipping.")
            continue

        img_h, img_w, _ = img.shape
        img_input = preprocess_onnx(img.copy())
        ort_inputs = {ort_session.get_inputs()[0].name: img_input}
        ort_outputs = ort_session.run(None, ort_inputs)
        detections = postprocess_onnx(ort_outputs[0], img_h, img_w)

        for x1, y1, x2, y2, confidence, class_id in detections:
            if confidence > 0.7 and class_id == 0:
                x1, y1 = max(0, x1), max(0, y1)
                x2, y2 = min(img_w, x2), min(img_h, y2)

                if x2 <= x1 or y2 <= y1:
                    print(f"Invalid bounding box for {filename}: ({x1}, {y1}, {x2}, {y2})")
                    continue

                face = img[y1:y2, x1:x2].copy()
                embedding = extract_embedding(face)

                if embedding is not None:
                    name = os.path.splitext(filename)[0]
                    faiss_index.add(embedding)
                    embedding_db[name] = embedding
                    print(f"Stored: {name} (Confidence: {confidence:.2f})")

    with open("face_index.pkl", "wb") as f:
        pickle.dump((embedding_db, faiss_index), f)
    print(f"Database updated. Total faces: {faiss_index.ntotal}")

def log_to_excel(name, timestamp):
    """Log the recognized name to an Excel file with a new sheet for each date"""
    date_str = timestamp.strftime("%Y-%m-%d")
    time_str = timestamp.strftime("%H:%M:%S")
    data = {"Name": [name], "Time": [time_str]}
    df = pd.DataFrame(data)
    
    excel_file = "attendance.xlsx"
    max_retries = 3
    retry_delay = 1  # seconds

    for attempt in range(max_retries):
        try:
            # Load or create the workbook
            if os.path.exists(excel_file):
                try:
                    workbook = openpyxl.load_workbook(excel_file)
                except InvalidFileException:
                    print(f"Excel file {excel_file} is corrupted. Creating a new one.")
                    workbook = openpyxl.Workbook()
                    workbook.save(excel_file)  # Save immediately to ensure file exists
            else:
                print(f"Creating new Excel file: {excel_file}")
                workbook = openpyxl.Workbook()
                workbook.save(excel_file)  # Save immediately to ensure file exists

            # Load existing sheet data or create new sheet
            if date_str in workbook.sheetnames:
                existing_df = pd.read_excel(excel_file, sheet_name=date_str)
                if "Name" not in existing_df.columns:
                    existing_df = pd.DataFrame(columns=["Name", "Time"])
                updated_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                updated_df = df

            # Write to the Excel file
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                updated_df.to_excel(writer, sheet_name=date_str, index=False)
            
            print(f"Logged to Excel: {name} at {date_str} {time_str} (Sheet: {date_str})")
            break
        except PermissionError:
            print(f"Permission denied for {excel_file}. Retrying in {retry_delay} seconds... ({attempt + 1}/{max_retries})")
            time.sleep(retry_delay)
        except Exception as e:
            print(f"Error writing to Excel: {e}")
            if attempt == max_retries - 1:
                print(f"Failed to write to {excel_file} after {max_retries} attempts. Saving backup.")
                backup_file = f"attendance_backup_{date_str}.xlsx"
                with pd.ExcelWriter(backup_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=date_str, index=False)
                print(f"Saved to backup: {backup_file}")
            time.sleep(retry_delay)
    else:
        print(f"Failed to write to {excel_file} after {max_retries} attempts.")

def recognize_live():
    """Recognizes faces in live video from webcam with Excel logging"""
    cap = cv2.VideoCapture(1)  # Open video file
    # cap = cv2.VideoCapture("D:/retinaface/recording1.mp4")  # Open video file

    if not cap.isOpened():
        print("Error: Could not open video.")
        return

    # Load existing names from Excel for each date to avoid duplicates on restart
    excel_file = "attendance.xlsx"
    if os.path.exists(excel_file):
        try:
            with pd.ExcelFile(excel_file) as xls:
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    if "Name" in df.columns:
                        logged_names[sheet_name] = set(df["Name"].unique())
                    else:
                        logged_names[sheet_name] = set()
                        df = pd.DataFrame(columns=["Name", "Time"])
                        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                        print(f"Initialized 'Name' column for sheet {sheet_name}")
        except InvalidFileException:
            print(f"Excel file {excel_file} is corrupted. Starting with a new file.")
            os.remove(excel_file)  # Remove corrupted file

    last_checked_date = None  # To track date changes during 24/7 operation

    while True:
        ret, frame = cap.read()
        if not ret:
            print("End of video or error capturing frame.")
            break

        img_h, img_w, _ = frame.shape
        img_input = preprocess_onnx(frame.copy())
        ort_inputs = {ort_session.get_inputs()[0].name: img_input}
        ort_outputs = ort_session.run(None, ort_inputs)
        detections = postprocess_onnx(ort_outputs[0], img_h, img_w)

        current_time = datetime.now()
        current_date_str = current_time.strftime("%Y-%m-%d")

        # Check if date has changed during execution
        if last_checked_date != current_date_str:
            if current_date_str not in logged_names:
                logged_names[current_date_str] = set()
            last_checked_date = current_date_str
            print(f"Date changed to {current_date_str}. Switching to new sheet.")

        # Check and remove names not detected within 2 seconds
        names_to_remove = []
        for name, data in list(detection_tracker.items()):
            if data["times"] and (current_time - data["times"][-1]) > timedelta(seconds=2) and data["count"] == 1:
                names_to_remove.append(name)
        for name in names_to_remove:
            del detection_tracker[name]
            print(f"Removed {name} from detection_tracker due to no detection in 2 seconds.")

        for x1, y1, x2, y2, confidence, class_id in detections:
            if confidence > 0.4 and class_id == 0:
                x1, y1 = max(0, x1), max(0, y1)
                x2, y2 = min(img_w, x2), min(img_h, y2)

                if x2 <= x1 or y2 <= y1:
                    print(f"Invalid bounding box: ({x1}, {y1}, {x2}, {y2})")
                    continue

                face = frame[y1:y2, x1:x2].copy()
                # if not is_face_straight(face):
                #     print("Skipping side-posed face.")
                #     continue

                aligned_face = align_face(face)
                aligned_face = cv2.resize(aligned_face, (112, 112))
                embedding = extract_embedding(aligned_face)

                if embedding is not None and faiss_index.ntotal > 0:
                    distances, indices = faiss_index.search(embedding, k=1)
                    closest_index = indices[0][0]
                    distance = distances[0][0]

                    name = list(embedding_db.keys())[closest_index] if distance < 0.4 else "Unknown"
                    if name.find("_") != -1:
                        name = name.split("_")[0]

                    print(f"Recognized: {name} (Distance: {distance:.2f}, Confidence: {confidence:.2f})")

                    # Update detection tracker
                    for x, y in detection_tracker.items():
                        print("detection_tracker -", x, y)
                    if name != "Unknown":
                        if name not in detection_tracker:
                            detection_tracker[name] = {"count": 0, "times": []}
                        
                        detection_tracker[name]["times"].append(current_time)
                        detection_tracker[name]["count"] += 1

                        # Clean up old timestamps (older than 2 seconds)
                        two_sec_ago = current_time - timedelta(seconds=2)
                        detection_tracker[name]["times"] = [t for t in detection_tracker[name]["times"] if t > two_sec_ago]
                        detection_tracker[name]["count"] = len(detection_tracker[name]["times"])

                        # Log to Excel if detected > 3 times in 2 seconds and not already logged for this date
                        if detection_tracker[name]["count"] > 3 and name not in logged_names[current_date_str]:
                            log_to_excel(name, current_time)
                            logged_names[current_date_str].add(name)  # Mark name as logged for this date
                            detection_tracker[name] = {"count": 0, "times": []}  # Reset tracker

                    cv2.putText(frame, f"{name} ({distance:.2f})", (x1, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)

        cv2.imshow("Live Face Recognition", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    # Train on your images (uncomment if needed)
    store_embeddings("images/Yashvee")
    store_embeddings("images/Ayush")
    # store_embeddings("Database/Anish")
    # store_embeddings("Database/Smit")
    recognize_live()